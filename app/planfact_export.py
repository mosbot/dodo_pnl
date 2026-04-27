"""Парсер отчёта «Прибыли и убытки по проектам» из ПланФакт (.xlsx).

Лист содержит дерево статей, заданное отступами в колонке A (4 пробела на
уровень). Помимо реальных статей в плоском виде встречаются «расчётные
строки» (Маржинальная прибыль, EBITDA, Чистая прибыль и т.п.), которые
не являются категориями учёта — мы их распознаём по фиксированному списку
имён и в шаблон не сохраняем.

Возвращаемая структура — плоский список узлов с parent_idx / depth /
path / is_leaf / is_calc / pnl_code (auto). Достаточно для отрисовки дерева
и сохранения в SQLite.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl

from .pnl import classify_category


# Названия строк, которые в экспорте идут на верхнем уровне, но являются
# не статьями, а расчётными показателями. Не сохраняем как статьи учёта,
# но сохраняем как узлы дерева — на дашборде они показываются как «—».
#
# ВАЖНО: «Амортизация», «Проценты по кредитам и займам», «Налог на прибыль (доходы)»,
# «Дивиденды» — это РЕАЛЬНЫЕ статьи P&L, у которых есть категории и обороты в учёте,
# а не расчётные метрики. Их в этот список НЕ добавляем, иначе на дашборде они
# теряют значения и показывают «—».
CALC_NAMES: set[str] = {
    "Маржинальная прибыль",
    "Маржинальность",
    "Операционная прибыль",
    "Операционная рентабельность",
    "EBITDA",
    "Рентабельность по EBITDA",
    "Чистая прибыль (убыток)",
    "Рентабельность чистой прибыли",
    "Нераспределенная прибыль",
}

# Топ-уровневые «секции» PnL — по ним определяется op_type/activity для
# симуляции classify_category() (в экспорте этой инфы нет, она только в API).
# В этот список включены и топ-уровневые «листовые» статьи, которые в xlsx
# идут плоско (без обёртки «Расходы / …»), но являются реальными статьями P&L
# с категориями и оборотами в учёте: «Амортизация», «Проценты по кредитам и
# займам», «Налог на прибыль (доходы)», «Дивиденды».
SECTION_SEMANTICS: dict[str, tuple[str, str]] = {
    "Выручка": ("Income", "Operating"),
    "Прочие доходы": ("Income", "Operating"),
    "Переменные расходы": ("Outcome", "Operating"),
    "Постоянные расходы": ("Outcome", "Operating"),
    "Прочие расходы": ("Outcome", "Operating"),
    # Below-the-line — финансовая/налоговая деятельность, но для целей шаблона
    # достаточно отнести к Outcome, чтобы classify_category() выбрал INTEREST/TAX/DIVIDENDS.
    "Амортизация": ("Outcome", "Operating"),
    "Проценты по кредитам и займам": ("Outcome", "Finance"),
    "Налог на прибыль (доходы)": ("Outcome", "Operating"),
    "Дивиденды": ("Outcome", "Capital"),
}


class ExportParseError(Exception):
    """Ошибка разбора файла экспорта."""


def _indent_depth(value: Any) -> int:
    """4 пробела (или НБ-пробела) = 1 уровень."""
    if value is None:
        return 0
    s = str(value).replace("\xa0", " ").replace("\t", "    ")
    leading = len(s) - len(s.lstrip(" "))
    return leading // 4


def _root_section_title(rows: list[dict], idx: int) -> str | None:
    """Поднимаемся к корню — название топ-уровневой секции."""
    cur = rows[idx]
    while cur["parent_idx"] is not None:
        cur = rows[cur["parent_idx"]]
    return None if cur["is_calc"] else cur["title"]


def parse_pnl_export(file_bytes: bytes) -> dict:
    """Разобрать xlsx-экспорт и вернуть список узлов.

    Возвращает {"nodes": [...], "warnings": [...]}, где node:
      - title: имя статьи
      - depth: уровень вложенности (0 = top)
      - parent_idx: индекс предка в массиве (None для top-level)
      - path: list[str] — путь от корня
      - is_leaf: True если нет детей в экспорте
      - is_calc: True если это расчётная строка (в шаблон не сохраняется)
      - pnl_code: предложенный код (UC/LC/DC/RENT/MARKETING/...) или None
      - sort_order: исходный порядок строк в файле (1..N)
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ExportParseError(f"Не удалось открыть xlsx: {e}") from e

    ws = wb.active
    if ws is None:
        raise ExportParseError("В файле нет ни одного листа.")

    # Шапка отчёта занимает первые ~7 строк, в строке 7 — заголовок «Статьи учёта».
    # Найдём её динамически: первая строка, где в колонке A написано ровно
    # «Статьи учета» / «Статьи учёта». Данные начинаются со следующей строки.
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip().lower().replace("ё", "е")
        if s == "статьи учета":
            header_row = r
            break
    if header_row is None:
        raise ExportParseError(
            "Не найдена строка заголовка «Статьи учёта». "
            "Проверь, что это экспорт «Отчёт о прибылях и убытках по проектам» из ПланФакт."
        )

    # Собираем плоский список статей.
    rows: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None:
            continue
        title = str(raw).replace("\xa0", " ").strip()
        if not title:
            continue
        depth = _indent_depth(raw)
        rows.append({
            "title": title,
            "depth": depth,
            "is_calc": title in CALC_NAMES,
        })

    # Реконструируем parent_idx по стеку индексов.
    stack: list[tuple[int, int]] = []  # (depth, idx_in_rows)
    for i, row in enumerate(rows):
        while stack and stack[-1][0] >= row["depth"]:
            stack.pop()
        row["parent_idx"] = stack[-1][1] if stack else None
        if stack:
            row["path"] = rows[stack[-1][1]]["path"] + [row["title"]]
        else:
            row["path"] = [row["title"]]
        stack.append((row["depth"], i))

    # Помечаем листья.
    parent_set = {r["parent_idx"] for r in rows if r["parent_idx"] is not None}
    for i, r in enumerate(rows):
        r["is_leaf"] = i not in parent_set

    # Авто-классификация в pnl_code: симулируем info-словарь, как в _build_category_index().
    warnings: list[str] = []
    for i, r in enumerate(rows):
        if r["is_calc"]:
            r["pnl_code"] = None
            continue
        section = _root_section_title(rows, i)
        if section is None:
            r["pnl_code"] = None
            continue
        op_type, activity = SECTION_SEMANTICS.get(section, (None, None))
        if op_type is None:
            warnings.append(f"Неизвестный корневой раздел: «{section}» — узел «{r['title']}» не классифицирован.")
            r["pnl_code"] = None
            continue
        info = {
            "title": r["title"],
            "path": r["path"],
            "path_str": " / ".join(r["path"]).lower(),
            "op_type": op_type,
            "activity_type": activity,
            "outcome_class": None,
        }
        r["pnl_code"] = classify_category(info)

    # sort_order — для сохранения исходного порядка после round-trip через БД.
    for i, r in enumerate(rows, start=1):
        r["sort_order"] = i

    return {
        "nodes": rows,
        "warnings": warnings,
        "leaf_count": sum(1 for r in rows if r["is_leaf"] and not r["is_calc"]),
        "calc_count": sum(1 for r in rows if r["is_calc"]),
        "total": len(rows),
    }
