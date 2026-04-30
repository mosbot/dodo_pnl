"""Генерация xlsx-выгрузок для P&L-детализации и drill-down операций.

Используется openpyxl (уже идёт через requirements). Выгрузка повторяет
структуру UI:
- Месяц: Статья / [проекты] / Итого
- Период: Статья / [месяцы] / Итого
- Drill-down: Дата / Статья / Проект / Контрагент / Комментарий / Сумма

Финансовые ячейки получают numFormat «#,##0» (или «#,##0;[Red]-#,##0»
для сумм, чтобы отрицательные подсвечивались). Иерархия в детализации
передаётся через отступ в первой колонке.
"""
from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL = PatternFill("solid", fgColor="F3F4F6")
_TOTAL_FILL = PatternFill("solid", fgColor="EFF6FF")
_THIN = Side(style="thin", color="E5E7EB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FMT_RUB = '#,##0;[Red]-#,##0'
_FMT_RUB_BOLD = _FMT_RUB
_FMT_PCT = '0.0%'


def _ru_month(month_key: str) -> str:
    """'2026-04' → 'Апрель 2026 г.'."""
    months_ru = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    try:
        y, m = month_key.split("-")
        return f"{months_ru[int(m) - 1]} {y} г."
    except Exception:
        return month_key


def _safe_filename(s: str) -> str:
    """Убирает кириллицу/опасные символы для Content-Disposition."""
    return "".join(c if c.isalnum() or c in "._-" else "_" for c in s)


def render_pnl_xlsx(
    *,
    pnl: dict,
    project_names: dict[str, str],
    period_label: str,
    selected_project_names: list[str],
    method: str,
    is_period_mode: bool,
) -> bytes:
    """Сгенерировать xlsx из dict-результата /api/pnl.

    pnl — результат _build_pnl_for_period + monthly breakdown (если Период).
    project_names — карта project_id → display_name (для шапки в Месяц-режиме).
    period_label — человекочитаемый период («Апрель 2026 г.» или «Март – Апрель 2026»).
    selected_project_names — список имён выбранных пиццерий для информационной шапки.
    is_period_mode — True если в xlsx нужны помесячные колонки.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"

    template_lines = pnl.get("template_lines") or []
    lines = pnl.get("lines") or []
    monthly = pnl.get("monthly") or {}
    months = pnl.get("months_in_range") or []
    projects = pnl.get("projects") or []

    # === Шапка отчёта (4 строки) ===
    ws["A1"] = "P&L · Детализация"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {period_label}"
    ws["A3"] = f"Метод: {'начисление' if method == 'accrual' else 'кассовый'}"
    ws["A4"] = "Проекты: " + (", ".join(selected_project_names) if selected_project_names else "—")
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

    # === Заголовки колонок ===
    header_row = 6
    headers: list[str] = ["Статья"]
    if is_period_mode:
        for m in months:
            headers.append(_ru_month(m))
    else:
        for p in projects:
            headers.append(p.get("name") or str(p.get("id")))
    headers.append("Итого")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # Ширина колонок
    ws.column_dimensions["A"].width = 48  # Статья
    n_data_cols = len(headers) - 1
    for i in range(n_data_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    # === Данные ===
    row = header_row + 1
    if template_lines:
        # Иерархия — отступы через level (depth). pct_row для % метрик.
        # build_pnl уже отдаёт template_lines плоским списком.
        for n in template_lines:
            depth = n.get("depth", 0) or 0
            is_calc = bool(n.get("is_calc"))
            is_header = depth == 0 and not is_calc
            display_kind = n.get("display_kind")
            is_pct_row = display_kind == "pct"

            indent = "    " * depth
            ws.cell(row=row, column=1, value=indent + (n.get("title") or ""))

            # Стилизация
            if is_header:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = _SECTION_FILL
                    ws.cell(row=row, column=c).font = Font(bold=True)
            elif is_calc:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = _TOTAL_FILL
                    ws.cell(row=row, column=c).font = Font(bold=True)

            # Значения по колонкам
            if is_period_mode:
                for col_i, m in enumerate(months, start=2):
                    by_node = (monthly.get(m) or {}).get("by_node") or {}
                    amt = by_node.get(str(n.get("id")))
                    cell = ws.cell(row=row, column=col_i)
                    if is_pct_row:
                        cell.value = None  # помесячный % не считаем
                    elif amt is not None:
                        cell.value = amt
                        cell.number_format = _FMT_RUB
            else:
                for col_i, p in enumerate(projects, start=2):
                    proj = (n.get("projects") or {}).get(p["id"]) or {}
                    cell = ws.cell(row=row, column=col_i)
                    if is_pct_row:
                        pct = proj.get("pct_of_revenue")
                        if pct is not None:
                            cell.value = pct
                            cell.number_format = _FMT_PCT
                    else:
                        amt = proj.get("amount")
                        if amt is not None:
                            cell.value = amt
                            cell.number_format = _FMT_RUB

            # Итого
            total = n.get("total") or {}
            total_col = len(headers)
            cell = ws.cell(row=row, column=total_col)
            if is_pct_row:
                pct = total.get("pct_of_revenue")
                if pct is not None:
                    cell.value = pct
                    cell.number_format = _FMT_PCT
            else:
                amt = total.get("amount")
                if amt is not None:
                    cell.value = amt
                    cell.number_format = _FMT_RUB_BOLD
            cell.font = Font(bold=True)

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = _BORDER

            row += 1
    else:
        # Fallback: агрегатная таблица (PNL_CODES). Используем lines.
        for ln in lines:
            kind = ln.get("kind", "detail")
            indent = ""
            if kind == "header":
                pass
            elif kind in ("summary", "final"):
                pass
            else:
                indent = "    "
            ws.cell(row=row, column=1, value=indent + (ln.get("label") or ""))
            if kind in ("header", "summary"):
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = _SECTION_FILL
                    ws.cell(row=row, column=c).font = Font(bold=True)
            elif kind == "final":
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = _TOTAL_FILL
                    ws.cell(row=row, column=c).font = Font(bold=True)

            if is_period_mode:
                for col_i, m in enumerate(months, start=2):
                    by_code = (monthly.get(m) or {}).get("by_code") or {}
                    amt = by_code.get(ln.get("code"))
                    cell = ws.cell(row=row, column=col_i)
                    if amt is not None:
                        cell.value = amt
                        cell.number_format = _FMT_RUB
            else:
                for col_i, p in enumerate(projects, start=2):
                    proj = (ln.get("projects") or {}).get(p["id"]) or {}
                    amt = proj.get("amount")
                    cell = ws.cell(row=row, column=col_i)
                    if amt is not None:
                        cell.value = amt
                        cell.number_format = _FMT_RUB

            total = ln.get("total") or {}
            cell = ws.cell(row=row, column=len(headers))
            amt = total.get("amount")
            if amt is not None:
                cell.value = amt
                cell.number_format = _FMT_RUB_BOLD
            cell.font = Font(bold=True)
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = _BORDER
            row += 1

    # Заморозить шапку
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_operations_xlsx(
    *,
    items: list[dict],
    sum_value: float,
    period_label: str,
    project_label: str,
    category_label: str,
) -> bytes:
    """Сгенерировать xlsx со списком операций (drill-down)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    # Шапка
    ws["A1"] = "Операции"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {period_label}"
    ws["A3"] = f"Проект: {project_label}"
    ws["A4"] = f"Статья: {category_label}"
    ws["A5"] = f"Операций: {len(items)} · Сумма: {sum_value:,.2f}".replace(",", " ")
    ws["A5"].font = Font(bold=True)

    # Заголовки таблицы — на 7-й строке
    headers = ["Дата", "Статья", "Проект", "Контрагент", "Комментарий", "Сумма"]
    header_row = 7
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER

    widths = [14, 32, 24, 28, 38, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Данные
    row = header_row + 1
    for op in items:
        ws.cell(row=row, column=1, value=op.get("date"))
        ws.cell(row=row, column=2, value=op.get("category") or "")
        ws.cell(row=row, column=3, value=op.get("project") or "")
        ws.cell(row=row, column=4, value=op.get("contrAgent") or "")
        ws.cell(row=row, column=5, value=op.get("comment") or "")
        cell = ws.cell(row=row, column=6, value=op.get("value"))
        if isinstance(op.get("value"), (int, float)):
            cell.number_format = _FMT_RUB
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = _BORDER
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_filename(prefix: str, period_label: str) -> str:
    """Безопасное имя файла. ASCII-only, .xlsx."""
    base = f"{prefix}-{period_label}".lower()
    return _safe_filename(base) + ".xlsx"
